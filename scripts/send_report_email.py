import os
import smtplib
import ssl
import datetime
import certifi
from pathlib import Path
from email.message import EmailMessage
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[1]
REPORT = BASE / "output" / "report_concorsi_v2_arricchito.txt"
FALLBACK_REPORT = BASE / "output" / "report_concorsi_v2.txt"
XLSX = BASE / "output" / "concorsi_dirigenza_UTILI_V2.xlsx"
ENV = BASE / "config" / "mail.env"
LOG = BASE / "logs" / "mail_error.log"

TARGET_EVALS = {"CANDIDARSI", "STUDIARE_SERIAMENTE", "STUDIARE"}

def load_env_file(path):
    data = {}
    if not path.exists():
        return data
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        data[k.strip()] = v.strip()
    return data

def get_config():
    env_file = load_env_file(ENV)
    return {
        "SMTP_USER": os.environ.get("SMTP_USER") or env_file.get("SMTP_USER"),
        "SMTP_PASS": os.environ.get("SMTP_PASS") or env_file.get("SMTP_PASS"),
        "SMTP_TO": os.environ.get("SMTP_TO") or env_file.get("SMTP_TO"),
    }

def count_new_relevant():
    if not XLSX.exists():
        return 0

    wb = load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    if "valutazione_umanoide" not in idx or "stato_novita" not in idx:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        valutazione = row[idx["valutazione_umanoide"]]
        stato = row[idx["stato_novita"]]
        if valutazione in TARGET_EVALS and stato == "NUOVO":
            count += 1

    return count

try:
    LOG.parent.mkdir(parents=True, exist_ok=True)

    new_count = count_new_relevant()

    if new_count == 0:
        LOG.write_text("Nessuna mail inviata: nessun nuovo bando strategico\n", encoding="utf-8")
        raise SystemExit(0)

    report_file = REPORT if REPORT.exists() else FALLBACK_REPORT

    if not report_file.exists():
        raise FileNotFoundError("Nessun report disponibile")

    body = report_file.read_text(encoding="utf-8")

    cfg = get_config()
    smtp_user = cfg["SMTP_USER"]
    smtp_pass = cfg["SMTP_PASS"]
    smtp_to = cfg["SMTP_TO"]

    if not smtp_user or not smtp_pass or not smtp_to:
        raise RuntimeError("SMTP_USER, SMTP_PASS o SMTP_TO mancanti")

    today = datetime.date.today().isoformat()

    msg = EmailMessage()
    msg["Subject"] = f"Monitor concorsi - {new_count} nuovi bandi strategici - {today}"
    msg["From"] = smtp_user
    msg["To"] = smtp_to
    msg.set_content(body)

    context = ssl.create_default_context(cafile=certifi.where())

    with smtplib.SMTP_SSL("smtp.mail.me.com", 465, context=context) as server:
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)

    LOG.write_text(f"Mail inviata correttamente: {new_count} nuovi bandi strategici\n", encoding="utf-8")

except SystemExit:
    raise
except Exception as e:
    LOG.write_text(f"ERRORE invio mail SMTP: {e}\n", encoding="utf-8")
    raise
