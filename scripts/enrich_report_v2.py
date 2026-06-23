import re
from pathlib import Path

import sys
PROFILE_SCRIPTS = Path(__file__).resolve().parents[1] / "profilo_strategico" / "scripts"
if str(PROFILE_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(PROFILE_SCRIPTS))
from cv_selector_core import choose_cv

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

BASE = Path(__file__).resolve().parents[1]
XLSX = BASE / "output" / "concorsi_dirigenza_UTILI_V2.xlsx"
OUT = BASE / "output" / "report_concorsi_v2_arricchito.txt"

HOME = "Casamarciano / Napoli"

TARGET_EVALS = {"CANDIDARSI", "STUDIARE_SERIAMENTE", "STUDIARE"}

CITY_DISTANCE = {
    "napoli": ("MOLTO_FAVOREVOLE", "area Napoli/Campania"),
    "caserta": ("MOLTO_FAVOREVOLE", "Campania"),
    "salerno": ("FAVOREVOLE", "Campania"),
    "benevento": ("FAVOREVOLE", "Campania"),
    "avellino": ("FAVOREVOLE", "Campania"),
    "roma": ("GESTIBILE", "Roma/Lazio"),
    "bari": ("MEDIA", "Sud, distanza gestibile ma non quotidiana"),
    "trento": ("ALTA_PENALITA", "Nord, trasferimento probabile"),
    "trieste": ("ALTA_PENALITA", "Nord-est, trasferimento probabile"),
    "pavia": ("ALTA_PENALITA", "Nord, trasferimento probabile"),
    "padova": ("ALTA_PENALITA", "Nord-est, trasferimento probabile"),
    "genova": ("ALTA_PENALITA", "Nord-ovest, trasferimento probabile"),
    "bologna": ("ALTA_PENALITA", "Nord, trasferimento probabile"),
    "milano": ("ALTA_PENALITA", "Nord, trasferimento probabile"),
    "torino": ("ALTA_PENALITA", "Nord-ovest, trasferimento probabile"),
    "firenze": ("MEDIA_ALTA", "Centro-nord, trasferimento probabile"),
}

def clean(s):
    return re.sub(r"\s+", " ", s or "").strip()

def infer_distance(title, detail):
    title_l = title.lower()
    detail_l = detail.lower()

    # Prima il titolo: è molto più affidabile del dettaglio InPA, che può contenere footer/testi nazionali.
    for city, value in CITY_DISTANCE.items():
        if city in title_l:
            return value

    # Poi ricerca mirata nel dettaglio, evitando di farsi ingannare da Roma nei footer/testi istituzionali.
    sede_patterns = [
        "sede di lavoro", "sede", "ente di riferimento", "presso", "università degli studi di",
        "politecnico di", "comune di", "provincia di"
    ]
    for pattern in sede_patterns:
        pos = detail_l.find(pattern)
        if pos != -1:
            window = detail_l[pos:pos+500]
            for city, value in CITY_DISTANCE.items():
                if city in window:
                    return value

    return ("DA_VERIFICARE", "sede non riconosciuta automaticamente")

def prestige(text):
    t = text.lower()
    if "università" in t or "politecnico" in t or "ateneo" in t:
        if any(x in t for x in ["pavia", "trento", "trieste", "padova", "bologna", "milano", "torino", "firenze", "napoli"]):
            return "ALTO - Ateneo/Politecnico rilevante"
        return "MEDIO_ALTO - Ente universitario"
    if any(x in t for x in ["ministero", "presidenza del consiglio", "anvur", "mur", "indire", "invalsi", "cineca"]):
        return "ALTO - Ente nazionale/strategico"
    if "comune" in t or "provincia" in t or "unione" in t:
        return "MEDIO - Ente territoriale"
    return "DA_VERIFICARE"

def why_coherent(title, detail):
    t = (title + " " + detail).lower()
    reasons = []

    if "comunicazione" in t or "relazioni esterne" in t:
        reasons.append("coerente con esperienza in comunicazione istituzionale, web, social e relazioni esterne")
    if "terza missione" in t or "public engagement" in t:
        reasons.append("coerente con public engagement, terza missione e valorizzazione istituzionale")
    if "sistemi informativi" in t or "trasformazione digitale" in t or "digitalizzazione" in t or "innovazione" in t:
        reasons.append("coerente con traiettoria comunicazione digitale / governance dei servizi digitali")
    if "area dei funzionari" in t or "elevata qualificazione" in t:
        reasons.append("coerente come passaggio ponte verso EQ, responsabilità organizzativa e crescita manageriale")
    if "dirigente" in t or "dirigenziale" in t:
        reasons.append("coerente come target alto per studio di profili dirigenziali e requisiti futuri")

    if not reasons:
        reasons.append("coerenza da verificare leggendo bando completo")

    return "; ".join(reasons)

def study_topics(title, detail):
    t = (title + " " + detail).lower()
    topics = {"diritto amministrativo", "pubblico impiego", "procedimento amministrativo", "trasparenza e anticorruzione"}

    if "università" in t or "ateneo" in t or "politecnico" in t:
        topics.update(["ordinamento universitario", "CCNL università", "governance degli atenei"])
    if "comunicazione" in t or "relazioni esterne" in t:
        topics.update(["comunicazione pubblica", "legge 150/2000", "accessibilità digitale", "social media policy PA"])
    if "terza missione" in t or "public engagement" in t:
        topics.update(["terza missione", "public engagement", "valutazione ANVUR", "rapporti con territorio e stakeholder"])
    if "sistemi informativi" in t or "digitalizzazione" in t or "trasformazione digitale" in t or "innovazione" in t:
        topics.update(["CAD", "transizione digitale", "cybersecurity PA", "project management ICT", "servizi digitali"])
    if "dirigente" in t or "dirigenziale" in t:
        topics.update(["management pubblico", "performance", "valore pubblico", "responsabilità dirigenziale", "contabilità pubblica base"])

    return ", ".join(sorted(topics))

def final_priority(eval_label, distance_label, prestige_label):
    score = 0

    if eval_label == "CANDIDARSI":
        score += 50
    elif eval_label == "STUDIARE_SERIAMENTE":
        score += 40
    elif eval_label == "STUDIARE":
        score += 30

    if distance_label in ["MOLTO_FAVOREVOLE", "FAVOREVOLE"]:
        score += 20
    elif distance_label == "GESTIBILE":
        score += 10
    elif distance_label == "MEDIA":
        score += 5
    elif distance_label in ["ALTA_PENALITA", "MEDIA_ALTA"]:
        score -= 10

    if prestige_label.startswith("ALTO"):
        score += 20
    elif prestige_label.startswith("MEDIO_ALTO"):
        score += 12
    elif prestige_label.startswith("MEDIO"):
        score += 5

    if score >= 75:
        return "PRIORITA_MASSIMA"
    if score >= 55:
        return "PRIORITA_ALTA"
    if score >= 35:
        return "PRIORITA_MEDIA"
    return "PRIORITA_BASSA"

def load_selected_rows():
    wb = load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {h: row[i] for h, i in idx.items()}
        if data.get("valutazione_umanoide") in TARGET_EVALS:
            rows.append(data)
    return rows


def metadata_quality(meta):
    ente = meta.get("ente", "DA_VERIFICARE")
    sede = meta.get("sede", "DA_VERIFICARE")
    scadenza = meta.get("scadenza", "DA_VERIFICARE")

    parts = []

    if ente != "DA_VERIFICARE":
        parts.append("ente: estratto")
    else:
        parts.append("ente: da verificare")

    if sede != "DA_VERIFICARE":
        parts.append("sede: estratta")
    else:
        parts.append("sede: da verificare")

    if scadenza != "DA_VERIFICARE":
        parts.append("scadenza: estratta")
    else:
        parts.append("scadenza: da verificare")

    return "; ".join(parts)

def practical_action(eval_label, dist_label, classe):
    eval_label = str(eval_label or "")
    dist_label = str(dist_label or "")
    classe = str(classe or "")

    if eval_label == "CANDIDARSI":
        if "ALTA_PENALITA" in dist_label:
            return "Valutare candidatura solo se profilo molto coerente; verificare sede, smart working, mobilità e sostenibilità logistica."
        return "Aprire il bando, verificare requisiti, preparare candidatura e adattare CV target."

    if eval_label == "STUDIARE_SERIAMENTE":
        if classe == "DIRIGENTE":
            return "Scaricare bando e requisiti; usarlo come benchmark per profilo dirigenziale, prove e gap da colmare."
        return "Studiare requisiti e prove; valutare candidatura solo dopo verifica logistica e requisiti."

    if eval_label == "STUDIARE":
        return "Archiviare come caso studio; estrarre materie, requisiti e lessico utile per preparazione."

    return "Monitorare senza azione immediata."

def crop_field(value):
    value = clean(value or "")
    stop_words = [
        " Questo sito", " Area geografica:", " Valutazione:", " Stato:",
        " Data apertura", " Data chiusura", " Numero posti",
        " Tipologia", " Candidati"
    ]
    for stop in stop_words:
        if stop in value:
            value = value.split(stop, 1)[0]
    return value.strip(" :-–—")[:140] or "DA_VERIFICARE"

def extract_first(patterns, text):
    import re
    text = clean(text or "")
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            value = crop_field(m.group(1))
            if value and value != "DA_VERIFICARE":
                return value
    return "DA_VERIFICARE"

def extract_metadata(title, detail):
    import re

    full = clean((title or "") + " " + (detail or ""))

    ente = extract_first([
        r"(Università degli Studi di [A-ZÀ-Üa-zà-ü'’ ]+)",
        r"(Università di [A-ZÀ-Üa-zà-ü'’ ]+)",
        r"(Politecnico di [A-ZÀ-Üa-zà-ü'’ ]+)",
        r"(?:Ente|Amministrazione|Ateneo)\s*[:\-–]\s*([^\n\r]+)",
        r"presso\s+(l['’]Università[^,\n\r\.]+)",
    ], full)

    sede = extract_first([
        r"\(([A-ZÀ-Üa-zà-ü'’ ]{3,60})\)\s+Numero posti",
        r"(?:Sede di lavoro|Luogo di lavoro|Comune)\s*[:\-–]\s*([^\n\r]+)",
        r"(?:presso la sede di|sede presso)\s+([^,\n\r]+)",
    ], full)

    scadenza = extract_first([
        r"Data chiusura candidature\s*[:\-–]?\s*([0-9]{1,2}\s+[A-ZÀ-Üa-zà-ü]+\s+[0-9]{4}(?:\s+[0-9]{1,2}[:\.][0-9]{2})?)",
        r"(?:Scadenza|Data scadenza|Termine presentazione domande|Termine per la presentazione delle domande)\s*[:\-–]\s*([0-9]{1,2}[\/\.-][0-9]{1,2}[\/\.-][0-9]{4}(?:\s*[0-9]{1,2}[:\.][0-9]{2})?)",
        r"(?:entro il|scade il|fino al)\s*([0-9]{1,2}[\/\.-][0-9]{1,2}[\/\.-][0-9]{4}(?:\s*[0-9]{1,2}[:\.][0-9]{2})?)",
    ], full)

    return {
        "ente": ente,
        "sede": sede,
        "scadenza": scadenza,
    }

def main():
    rows = load_selected_rows()
    lines = []
    lines.append("REPORT CONCORSI ARRICCHITO")
    lines.append(f"Base logistica: {HOME}")
    lines.append("=" * 70)
    lines.append("")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for n, r in enumerate(rows, 1):
            title = clean(r.get("titolo"))
            link = r.get("link")
            eval_label = r.get("valutazione_umanoide")
            classe = r.get("classe")
            score = r.get("score")

            detail = ""
            try:
                page.goto(link, wait_until="networkidle", timeout=60000)
                detail = clean(page.locator("body").inner_text())[:7000]
            except Exception as e:
                detail = f"ERRORE LETTURA DETTAGLIO: {e}"

            dist_label, dist_note = infer_distance(title, detail)
            prestige_label = prestige(title + " " + detail)
            why = why_coherent(title, detail)
            study = study_topics(title, detail)
            priority = final_priority(eval_label, dist_label, prestige_label)
            meta = extract_metadata(title, detail)

            lines.append(f"{n}. {priority} | {eval_label} | {classe} | Score {score}")
            lines.append(title)
            lines.append("")
            lines.append(f"Ente: {meta['ente']}")
            lines.append(f"Sede reale: {meta['sede']}")
            lines.append(f"Scadenza: {meta['scadenza']}")
            lines.append(f"Fonte dati: {metadata_quality(meta)}")
            lines.append(f"Azione pratica: {practical_action(eval_label, dist_label, classe)}")
            lines.append("")
            lines.append(f"Perché coerente: {why}")
            lines.append(f"Cosa studiare: {study}")
            lines.append(f"Prestigio: {prestige_label}")
            lines.append(f"Distanza/logistica da {HOME}: {dist_label} - {dist_note}")
            cv_info = choose_cv(title)
            lines.append(f"CV consigliato: {cv_info['cv']}")
            lines.append("Progetti da evidenziare: " + "; ".join(cv_info["projects"]))
            lines.append("Gap da coprire: " + "; ".join(cv_info["gaps"]))
            lines.append(f"Link: {link}")
            lines.append("-" * 70)
            lines.append("")

        browser.close()

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Report arricchito creato: {OUT}")

if __name__ == "__main__":
    main()
