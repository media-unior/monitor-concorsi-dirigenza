import os
import smtplib
import ssl
from pathlib import Path
from email.message import EmailMessage
from datetime import datetime

import certifi
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[1]
OUTPUT = BASE / "output"
LOGS = BASE / "logs"

ARCHIVIO = OUTPUT / "concorsi_dirigenza_ARCHIVIO_COMPLETO.xlsx"
UTILI_V2 = OUTPUT / "concorsi_dirigenza_UTILI_V2.xlsx"
REPORT = OUTPUT / "report_concorsi_v2_arricchito.txt"

def env(name):
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Variabile mancante: {name}")
    return value

def count_rows(path):
    if not path.exists():
        return 0
    wb = load_workbook(path)
    ws = wb.active
    return max(ws.max_row - 1, 0)

def read_tail(path, n=20):
    if not path.exists():
        return "Log non presente."
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    return "\n".join(lines[-n:]) if lines else "Log vuoto."

def main():
    smtp_user = env("SMTP_USER")
    smtp_pass = env("SMTP_PASS")
    smtp_to = env("SMTP_TO")

    archivio_count = count_rows(ARCHIVIO)
    utili_count = count_rows(UTILI_V2)

    body = []
    body.append("REPORT SETTIMANALE MONITOR CONCORSI")
    body.append("")
    body.append(f"Data controllo: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    body.append("")
    body.append(f"Bandi in archivio completo: {archivio_count}")
    body.append(f"Bandi strategici V2 rilevati: {utili_count}")
    body.append("")
    body.append("Stato: monitor cloud operativo.")
    body.append("")
    body.append("Nota: le mail ordinarie vengono inviate solo in presenza di nuovi bandi strategici.")
    body.append("")
    body.append("Ultime righe log monitor V2:")
    body.append(read_tail(LOGS / "monitor_v2.log", 20))

    if REPORT.exists():
        body.append("")
        body.append("Ultimo report arricchito disponibile negli artifact GitHub Actions.")

    msg = EmailMessage()
    msg["Subject"] = "Monitor concorsi - report settimanale"
    msg["From"] = smtp_user
    msg["To"] = smtp_to
    msg.set_content("\n".join(body))

    context = ssl.create_default_context(cafile=certifi.where())

    with smtplib.SMTP("smtp.mail.me.com", 587, timeout=30) as server:
        server.ehlo()
        server.starttls(context=context)
        server.ehlo()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)

    print("Report settimanale inviato")

if __name__ == "__main__":
    main()
