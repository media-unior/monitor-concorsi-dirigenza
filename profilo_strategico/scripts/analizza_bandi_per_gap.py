from pathlib import Path
from openpyxl import load_workbook
import csv
import re

ROOT = Path(__file__).resolve().parents[2]
BASE = ROOT / "profilo_strategico"
XLSX = ROOT / "output" / "concorsi_dirigenza_UTILI_V2.xlsx"
OUT = BASE / "output" / "analisi_gap_da_bandi.md"
GAP_OUT = BASE / "data" / "gap_da_bandi.csv"

AREE = {
    "diritto amministrativo": ["procedimento", "amministrativo", "trasparenza", "anticorruzione", "accesso"],
    "pubblico impiego": ["personale", "ccnl", "funzionari", "elevata qualificazione", "dirigente", "mobilità"],
    "governance universitaria": ["università", "ateneo", "ricerca", "terza missione", "governance"],
    "comunicazione pubblica": ["comunicazione", "media", "relazioni esterne", "web", "social"],
    "digitale": ["sistemi informativi", "digitale", "digitalizzazione", "innovazione", "cad", "portale"],
    "management pubblico": ["dirigente", "programmazione", "performance", "organizzativi", "responsabilità"],
    "contabilità pubblica": ["bilancio", "tributi", "programmazione", "contabilità"],
    "terza missione": ["terza missione", "public engagement", "territorio", "stakeholder"],
}

TARGET = {"CANDIDARSI", "STUDIARE_SERIAMENTE", "STUDIARE"}

def norm(x):
    return str(x or "").lower()

def main():
    if not XLSX.exists():
        raise FileNotFoundError(f"File non trovato: {XLSX}")

    wb = load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        valutazione = row[idx.get("valutazione_umanoide")]
        if valutazione not in TARGET:
            continue

        titolo = row[idx.get("titolo")]
        match = row[idx.get("match")]
        classe = row[idx.get("classe")]
        compat = row[idx.get("compatibilita")]
        priorita = row[idx.get("priorita_biagio")]
        link = row[idx.get("link")]

        text = norm(titolo) + " " + norm(match)

        found = []
        for area, keys in AREE.items():
            if any(k in text for k in keys):
                found.append(area)

        rows.append({
            "titolo": titolo,
            "valutazione": valutazione,
            "classe": classe,
            "compatibilita": compat,
            "priorita": priorita,
            "aree": found,
            "link": link,
        })

    area_count = {}
    for r in rows:
        for a in r["aree"]:
            area_count[a] = area_count.get(a, 0) + 1

    lines = []
    lines.append("# Analisi gap da bandi reali")
    lines.append("")
    lines.append("Fonte: output/concorsi_dirigenza_UTILI_V2.xlsx")
    lines.append("")
    lines.append("## Frequenza aree competenza")
    lines.append("")
    for area, count in sorted(area_count.items(), key=lambda x: x[1], reverse=True):
        lines.append(f"- {area}: {count}")
    lines.append("")

    lines.append("## Bandi analizzati")
    lines.append("")
    for i, r in enumerate(rows, 1):
        lines.append(f"### {i}. {r['valutazione']} | {r['classe']} | {r['priorita']}")
        lines.append("")
        lines.append(str(r["titolo"]))
        lines.append("")
        lines.append("Aree competenza emerse:")
        for a in r["aree"]:
            lines.append(f"- {a}")
        lines.append("")
        lines.append(f"Link: {r['link']}")
        lines.append("")

    lines.append("## Indicazione operativa")
    lines.append("")
    lines.append("Le aree più frequenti devono guidare aggiornamento CV, piano studio, corsi brevi e scelta incarichi interni.")
    lines.append("")

    OUT.write_text("\n".join(lines), encoding="utf-8")

    with GAP_OUT.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["area", "frequenza", "azione"])
        for area, count in sorted(area_count.items(), key=lambda x: x[1], reverse=True):
            writer.writerow([area, count, "rafforzare se frequente nei bandi target"])

    print(f"Analisi creata: {OUT}")
    print(f"CSV creato: {GAP_OUT}")

if __name__ == "__main__":
    main()
