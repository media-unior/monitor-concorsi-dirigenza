import csv
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[2]
PROFILO = BASE / "profilo_strategico"
INPUT = BASE / "output" / "concorsi_dirigenza_UTILI_V2.xlsx"
REPORT = BASE / "output" / "report_concorsi_v2_arricchito.txt"
FOLLOWUP = PROFILO / "data" / "bandi_followup.csv"
OUT = PROFILO / "output" / "report_bandi_followup.md"

FIELDS = [
    "data_importazione",
    "titolo",
    "ente",
    "sede",
    "scadenza",
    "classe",
    "valutazione",
    "priorita",
    "score",
    "cv_consigliato",
    "link",
    "stato_followup",
    "data_primo_controllo",
    "note",
]

def read_existing():
    if not FOLLOWUP.exists():
        return [], set()

    with FOLLOWUP.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))

    links = {r.get("link", "").strip() for r in rows if r.get("link")}
    return rows, links

def write_rows(rows):
    FOLLOWUP.parent.mkdir(parents=True, exist_ok=True)
    with FOLLOWUP.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)

def parse_report_blocks():
    if not REPORT.exists():
        return {}

    text = REPORT.read_text(encoding="utf-8", errors="ignore")
    blocks = text.split("-" * 70)
    data = {}

    for block in blocks:
        link = ""
        item = {
            "ente": "DA_VERIFICARE",
            "sede": "DA_VERIFICARE",
            "scadenza": "DA_VERIFICARE",
            "cv_consigliato": "cv_strategico_sintetico.md",
            "priorita": "DA_VERIFICARE",
        }

        for line in block.splitlines():
            line = line.strip()
            if " | " in line and "Score" in line and "." in line[:5]:
                item["priorita"] = line.split(".", 1)[1].split("|", 1)[0].strip()
            elif line.startswith("Ente:"):
                item["ente"] = line.split(":", 1)[1].strip()
            elif line.startswith("Sede reale:"):
                item["sede"] = line.split(":", 1)[1].strip()
            elif line.startswith("Scadenza:"):
                item["scadenza"] = line.split(":", 1)[1].strip()
            elif line.startswith("CV consigliato:"):
                item["cv_consigliato"] = line.split(":", 1)[1].strip()
            elif line.startswith("Link:"):
                link = line.split(":", 1)[1].strip()

        if link:
            data[link] = item

    return data

def first_check_date(scadenza):
    if not scadenza or scadenza == "DA_VERIFICARE":
        return "DA_VERIFICARE"

    months = {
        "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
        "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
        "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
    }

    parts = scadenza.lower().split()
    try:
        day = int(parts[0])
        month = months[parts[1]]
        year = int(parts[2])
        d = datetime(year, month, day) + timedelta(days=20)
        return d.strftime("%Y-%m-%d")
    except Exception:
        return "DA_VERIFICARE"

def load_useful_calls():
    if not INPUT.exists():
        raise FileNotFoundError(f"File non trovato: {INPUT}")

    wb = load_workbook(INPUT)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, r))
        eval_label = str(item.get("valutazione_umanoide") or "")
        if eval_label not in {"CANDIDARSI", "STUDIARE_SERIAMENTE", "STUDIARE"}:
            continue
        rows.append(item)

    return rows

def generate_report(rows):
    lines = []
    lines.append("# Report follow-up bandi")
    lines.append("")
    lines.append(f"Totale bandi in follow-up: {len(rows)}")
    lines.append("")

    pending = [r for r in rows if r.get("stato_followup") == "DA_MONITORARE"]
    lines.append(f"Da monitorare: {len(pending)}")
    lines.append("")

    for i, r in enumerate(rows, 1):
        lines.append(f"## {i}. {r.get('titolo')}")
        lines.append("")
        lines.append(f"- Ente: {r.get('ente')}")
        lines.append(f"- Sede: {r.get('sede')}")
        lines.append(f"- Scadenza: {r.get('scadenza')}")
        lines.append(f"- Primo controllo esito: {r.get('data_primo_controllo')}")
        lines.append(f"- Valutazione: {r.get('valutazione')}")
        lines.append(f"- Priorità: {r.get('priorita')}")
        lines.append(f"- CV: {r.get('cv_consigliato')}")
        lines.append(f"- Stato: {r.get('stato_followup')}")
        lines.append(f"- Link: {r.get('link')}")
        lines.append("")

    OUT.write_text("\n".join(lines), encoding="utf-8")

def main():
    existing_rows, existing_links = read_existing()
    report_data = parse_report_blocks()
    useful = load_useful_calls()

    today = datetime.now().strftime("%Y-%m-%d")
    new_rows = []

    for item in useful:
        link = str(item.get("link") or "").strip()
        if not link or link in existing_links:
            continue

        meta = report_data.get(link, {})
        scadenza = meta.get("scadenza", "DA_VERIFICARE")

        new_rows.append({
            "data_importazione": today,
            "titolo": str(item.get("titolo") or "").strip(),
            "ente": meta.get("ente", str(item.get("ente") or "DA_VERIFICARE")),
            "sede": meta.get("sede", "DA_VERIFICARE"),
            "scadenza": scadenza,
            "classe": str(item.get("classe") or ""),
            "valutazione": str(item.get("valutazione_umanoide") or ""),
            "priorita": meta.get("priorita", str(item.get("priorita_biagio") or "")),
            "score": str(item.get("score") or ""),
            "cv_consigliato": meta.get("cv_consigliato", "cv_strategico_sintetico.md"),
            "link": link,
            "stato_followup": "DA_MONITORARE",
            "data_primo_controllo": first_check_date(scadenza),
            "note": "",
        })

    all_rows = existing_rows + new_rows
    write_rows(all_rows)
    generate_report(all_rows)

    print(f"Nuovi bandi aggiunti al follow-up: {len(new_rows)}")
    print(f"Totale bandi in follow-up: {len(all_rows)}")
    print(f"Report creato: {OUT}")

if __name__ == "__main__":
    main()
